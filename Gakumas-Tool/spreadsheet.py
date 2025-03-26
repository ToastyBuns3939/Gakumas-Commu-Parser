import openpyxl
import openpyxl.styles
from data_types import TranslationLine

column_headers = ("type", "name", "translated name", "text", "translated text")


def convert_to_string(data):
    if data is None:
        return ""
    else:
        return str(data)


def row_to_translation_line(data_row) -> TranslationLine:
    return TranslationLine(*(convert_to_string(a) for a in data_row))


def is_blank(row: tuple):
    return all(item is None for item in row)


def get_tl_lines_from_spreadsheet(
    spreadsheet_path: str, worksheet_name: str
) -> list[TranslationLine]:
    # Read data from workbook
    workbook = openpyxl.load_workbook(filename=spreadsheet_path)
    existing_rows = [
        row
        for row in workbook[worksheet_name].iter_rows(
            min_col=1, max_col=5, values_only=True
        )
        if not is_blank(row)
    ]

    # check it has the right column headers
    existing_column_headers = tuple(str(header) for header in existing_rows[0])
    if existing_column_headers != column_headers:
        raise Exception(
            f"Existing spreadsheet has incorrect column headers!"
            + f"Expected headers\n{",".join(column_headers)}\n"
            + f"but spreadsheet has headers\n{",".join(existing_column_headers)}"
        )

    return [row_to_translation_line(data_row) for data_row in existing_rows[1:]]


def write_tl_lines_to_spreadsheet(
    tl_lines: list[TranslationLine], output_path: str, worksheet_name: str
):
    # Create the spreadsheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if not worksheet:
        raise Exception("Workbook has no active worksheet!")
    worksheet.title = worksheet_name

    # Write data to the worksheet
    worksheet.append(column_headers)
    for index, tl_line in enumerate(tl_lines):
        worksheet.append(tl_line)
        # Adjust row heights automatically based on the content
        text_line_count = tl_line.text.count("\n") + 1
        translated_text_line_count = tl_line.translated_text.count("\n") + 1
        row_height = 15 * max(  # Assuming default row height is 15 units
            text_line_count, translated_text_line_count
        )
        worksheet.row_dimensions[index + 2].height = row_height

    # Set column widths
    worksheet.column_dimensions["A"].width = 15  # type column
    worksheet.column_dimensions["B"].width = 15  # name column
    worksheet.column_dimensions["C"].width = 15  # translated name column
    worksheet.column_dimensions["D"].width = 40  # text column
    worksheet.column_dimensions["E"].width = 40  # translated column

    # Define formats for cell alignment and text wrapping
    dialogue_alignment = openpyxl.styles.Alignment(
        horizontal="left",
        vertical="top",
        # Excel doesn't display multiple lines unless the
        # cell has text wrapping
        # wrap_text = True
        # However Google Sheets does display multiple lines fine
        # without text wrapping, which is preferrable
    )
    name_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    # Apply alignment formats
    for row in worksheet["A:C"]:
        for cell in row:
            cell.alignment = name_alignment
    for row in worksheet["D:E"]:
        for cell in row:
            cell.alignment = dialogue_alignment

    # Close the workbook to save the Excel file
    workbook.save(output_path)
