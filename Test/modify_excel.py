import openpyxl
import re
import os
import sys

def replace_ellipsis(text):
    """Rule: Replace '…' (U+2026 HORIZONTAL ELLIPSIS) with '...' (three full stops)."""
    return text.replace("…", "...")

def replace_tilde(text):
    """Rule: Replace '~' (U+007E TILDE) with '～' (U+FF5E FULLWIDTH TILDE)."""
    return text.replace("~", "～")

def normalize_dashes(text):
    """
    Consolidates dash-related transformations for general dash normalization:
    1. Removes spaces before -, ー, or — (but NOT before ――).
    2. Converts various double dash forms (-- , ーー, ——, ──) to '――' (U+2015 HORIZONTAL BAR).
    3. Extends existing '――' when immediately followed by another single dash (-, ー, or —).
    """
    # Rule 3: Remove space before -, ー, or —
    text = re.sub(r'\s+([-ー—])', r'\1', text)

    # Rules 4, 5, 6, 7: Convert various double dash forms to '――'
    text = text.replace('--', '――')   # Double ASCII hyphen
    text = text.replace('ーー', '――') # Double long dash (U+30FC KATAKANA-HIRAGANA PROLONGED SOUND MARK)
    text = text.replace('——', '――') # Double em dash (U+2014 EM DASH)
    text = text.replace('──', '――') # Double box-drawing hyphen (U+2500 BOX DRAWINGS LIGHT HORIZONTAL)

    # Rule 8: Extend horizontal bar
    text = re.sub(r'(――+)([-ー—])', r'\1――', text)
    return text

def convert_dash_with_exceptions(text):
    """
    Converts single hyphens (-, ー, —) to '――' (U+2015 HORIZONTAL BAR)
    with specific exceptions to preserve hyphens used in:
    1. Word-hyphen-word patterns (e.g., "well-being").
    2. Placeholders like {user}- or {username}-.
    3. HTML-like tags like </r>- followed by a word character.
    """
    def _convert_dash_with_exceptions_logic(match):
        dash = match.group(0) # The matched dash character (-, ー, or —)
        start_index = match.start()
        full_string = match.string # The entire string being processed

        # Exception 1: Hyphen strictly between two word characters (\w-\w)
        # Checks if there's a word character before AND after the dash.
        if start_index > 0 and start_index < len(full_string) - 1:
            pre_char = full_string[start_index - 1]
            post_char = full_string[start_index + 1]
            if re.match(r'\w', pre_char) and re.match(r'\w', post_char):
                return dash # Keep the hyphen as is

        # Exception 2: Hyphen immediately following {user} or {username} and followed by a word character
        # Checks for '{user}-word' or '{username}-word'
        if start_index >= len('{user}') and full_string[start_index - len('{user}'):start_index] == '{user}':
            if start_index < len(full_string) - 1 and re.match(r'\w', full_string[start_index + 1]):
                return dash # Keep the hyphen
        if start_index >= len('{username}') and full_string[start_index - len('{username}'):start_index] == '{username}':
            if start_index < len(full_string) - 1 and re.match(r'\w', full_string[start_index + 1]):
                return dash # Keep the hyphen

        # Exception 3: Hyphen immediately following </r> and followed by a word character
        # Checks for '</r>-word'
        if start_index >= len('</r>') and full_string[start_index - len('</r>'):start_index] == '</r>':
            if start_index < len(full_string) - 1 and re.match(r'\w', full_string[start_index + 1]):
                return dash # Keep the hyphen

        # If none of the exceptions match, convert the dash to '――'
        return '――'

    # Apply the custom conversion logic to all single '-', 'ー', or '—' characters.
    return re.sub(r'([-ー—])', _convert_dash_with_exceptions_logic, text)

def convert_em_to_italic_tags(text):
    """
    Rule: Replace all variations of opening <em...> tags (e.g., <em\=>, <em=>, <em>) with <i>,
    and closing </em> tags with </i>.
    """
    # Updated Regex:
    # r'<em(\\)?(\s*=\s*)?>'
    #   <em        - matches literal "<em"
    #   (\\)?      - matches an optional backslash (escaped with another backslash)
    #   (\s*=\s*)? - matches an optional group for "=" with surrounding spaces
    #   >          - matches literal ">"
    text = re.sub(r'<em(\\)?(\s*=\s*)?>', '<i>', text, flags=re.IGNORECASE)
    # Replace </em> with </i>.
    text = text.replace('</em>', '</i>')
    return text

# --- Mapping of Function Names to Function Objects ---

# This dictionary allows us to select functions by their string names
TRANSFORMATION_FUNCTIONS = {
    'replace_ellipsis': replace_ellipsis,
    'replace_tilde': replace_tilde,
    'normalize_dashes': normalize_dashes,
    'convert_dash_with_exceptions': convert_dash_with_exceptions,
    'convert_em_to_italic_tags': convert_em_to_italic_tags,
}

# --- Core Excel Modification Function ---

def modify_excel_column_forward(input_filepath, output_filepath, functions_to_apply):
    """
    Modifies cells in the 'translated text' column of an Excel file
    based on a list of specified transformation functions.
    Preserves original cell formatting.
    **Only saves the file if actual changes were detected** in the 'translated text' column.

    Args:
        input_filepath (str): The path to the input Excel file.
        output_filepath (str): The path where the modified Excel file will be saved.
        functions_to_apply (list): A list of function objects (from TRANSFORMATION_FUNCTIONS) to apply.
    """
    try:
        workbook = openpyxl.load_workbook(input_filepath)
        sheet = workbook.active # Work on the active sheet

        # Find the column index for 'translated text' based on header row
        header_row = sheet[1] # Assuming header is in the first row
        translated_text_col_index = -1
        for col_index, cell in enumerate(header_row):
            # Case-insensitive and whitespace-agnostic matching for the header
            if isinstance(cell.value, str) and cell.value.strip().lower() == 'translated text':
                translated_text_col_index = col_index + 1 # openpyxl columns are 1-indexed
                break

        if translated_text_col_index == -1:
            print(f"Warning: 'translated text' column not found in '{os.path.basename(input_filepath)}'. Skipping this file.")
            return # Exit if the column isn't found

        changes_detected = False # Flag to track if any cell value was actually modified

        # Iterate through rows, starting from the second row to skip the header.
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=translated_text_col_index)
            original_value = cell.value # Store the cell's original value

            # Only process if the cell contains a string
            if isinstance(original_value, str):
                modified_value = original_value

                # Apply each selected transformation function in the defined order
                for func in functions_to_apply:
                    modified_value = func(modified_value)

                # If the modified value is different from the original, update the cell
                if modified_value != original_value:
                    cell.value = modified_value
                    changes_detected = True # Set the flag to True

        # Save the workbook only if changes were detected
        if changes_detected:
            workbook.save(output_filepath)
            print(f"Successfully applied transformations to '{os.path.basename(input_filepath)}' and saved to '{os.path.basename(output_filepath)}'")
        else:
            print(f"No changes detected in '{os.path.basename(input_filepath)}' for the selected transformations. Skipping save.")

    except FileNotFoundError:
        print(f"Error: Input file not found at '{input_filepath}' during processing.")
    except Exception as e:
        print(f"An unexpected error occurred while processing '{os.path.basename(input_filepath)}': {e}")


# --- Script Execution Logic (main block) ---

if __name__ == "__main__":
    # Check for correct number of command-line arguments
    # Expected: script_name.py <input_folder> <output_folder> [optional_function_names...]
    if len(sys.argv) < 3:
        print("Usage: python modify_excel.py <input_folder_path> <output_folder_path> [function_name1] [function_name2]...")
        print("\nAvailable functions (case-sensitive):")
        for func_name in TRANSFORMATION_FUNCTIONS.keys():
            print(f"  - {func_name}")
        sys.exit(1) # Exit with an error code

    input_folder = sys.argv[1]
    output_folder = sys.argv[2]

    # Determine which functions to apply based on command-line arguments
    selected_function_names = sys.argv[3:]
    if not selected_function_names:
        # If no specific functions are provided, apply all available ones
        print("No specific functions selected. Applying all available transformations by default.")
        functions_to_run = list(TRANSFORMATION_FUNCTIONS.values())
    else:
        # Collect the actual function objects based on provided names
        functions_to_run = []
        invalid_functions = []
        for name in selected_function_names:
            if name in TRANSFORMATION_FUNCTIONS:
                functions_to_run.append(TRANSFORMATION_FUNCTIONS[name])
            else:
                invalid_functions.append(name)

        # Report any unrecognized function names
        if invalid_functions:
            print(f"Warning: The following function(s) were not recognized and will be skipped: {', '.join(invalid_functions)}")
            print("Please ensure function names are spelled correctly (they are case-sensitive).")

        # Exit if no valid functions were specified at all
        if not functions_to_run:
            print("No valid functions were selected for execution. Exiting.")
            sys.exit(1)

    # --- Pre-processing Checks ---

    # Ensure the output folder exists; create it if it doesn't
    if not os.path.exists(output_folder):
        try:
            os.makedirs(output_folder)
            print(f"Created output folder: '{output_folder}'")
        except OSError as e:
            print(f"Error: Could not create output folder '{output_folder}': {e}")
            sys.exit(1) # Exit if output folder cannot be created

    # Check if the input folder exists and is a directory
    if not os.path.isdir(input_folder):
        print(f"Error: The input path '{input_folder}' is not a valid directory.")
        print("Please check the path and try again.")
        sys.exit(1) # Exit if input path is invalid
    else:
        # --- Main Processing Loop ---
        print(f"Processing Excel files in '{input_folder}'...")
        # Display the names of the functions that will be applied for clarity
        print(f"Applying transformations: {[func.__name__ for func in functions_to_run]}")

        # Iterate through all files in the input folder
        for filename in os.listdir(input_folder):
            input_filepath = os.path.join(input_folder, filename)

            # Process only Excel files (specifically .xlsx format which openpyxl handles best)
            if os.path.isfile(input_filepath) and filename.endswith('.xlsx'):
                output_filepath = os.path.join(output_folder, filename)
                # Call the main modification function for each Excel file
                modify_excel_column_forward(input_filepath, output_filepath, functions_to_run)

        print("\nBatch conversion complete.")